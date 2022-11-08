from tkinter import *
from tkinter import filedialog
import openpyxl as xl
import datetime
from dateutil.relativedelta import relativedelta

#Criando o App
root = Tk()
root.title('Excel Sheet Creator')
root.geometry('300x280')
root.resizable(width=False, height=False)

#Definindo variáveis
variable = StringVar()
variable.set('Selecionar opção')
descricao = ''
lista_tipo = [
    "Veículos",
    "Móveis", 
    "Equipamentos", 
    "Imóveis", 
    "Instalações"
]

#Definindo functions
#Função para criar a planilha
def criar_planilha():

    #Definindo variáveis
    num = 1
    valor = send_entries()[0]
    parcelas = send_entries()[1]
    tipo = send_entries()[2]
    data = send_entries()[3]
    conta_debito = '3.01.01.08.01.11'
    conta_credito = send_entries()[4]
    descricao = send_entries()[5]
    valor_mensal = round(float(valor) / int(parcelas), 2)
     
    #Criando o Workbook e a Worksheet
    wb = xl.Workbook()
    ws = wb.active
    ws.title = 'Depreciação'

    #Definindo o nome das colunas
    ws['A1'] = 'DATA'
    ws['B1'] = 'DESCRIÇÃO'
    ws['C1'] = 'VALOR'
    ws['D1'] = 'CONTA DÉBITO'
    ws['E1'] = 'CONTA CRÉDITO'

    #Definindo a dimensão da coluna
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 36
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15

    #Definindo a data
    data_final = data
    formato_data = '%d/%m/%Y'
    data_formatada = datetime.datetime.strptime(data_final,formato_data) 
    
    #Definindo a coluna de data
    for row in range(2, int(parcelas) + 2):
        cell = ws.cell(row,1)
        cell.value = data_formatada.date() 
        #Incrementa um mês após colocar a data na planilha
        data_formatada = data_formatada + relativedelta(months=1) 
        
        #Esse If e Else foram criados para que consigamos acessar sempre último dia do mês
        if data_formatada.month == 12:
            data_formatada = data_formatada.replace(day=31)
        else:
            data_formatada = data_formatada.replace(month=data_formatada.month+1, day=1) - datetime.timedelta(days=1)

    #Definindo a coluna de descrição
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row,2)
        cell.value = descricao + str(num) + "/" + str(parcelas) 
        num += 1

    #Definindo a coluna de valor
    for row in range(2,ws.max_row + 1):
        cell = ws.cell(row,3)
        cell.value = valor_mensal  

    #Definindo a coluna de conta débito
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row,4)
        cell.value = conta_debito 

    #Definindo a coluna de conta crédito
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row,5)
        cell.value = conta_credito 

    #Criar o arquivo excel
    wb.save(str(save_spot) + '/nova_planilha.xlsx')

#Função do botão send
def send_entries():
    valor = entrada_valor.get()
    tipo = variable.get()
    if tipo == lista_tipo[0]:
        tipo = 0.2
        descricao = 'Depreciação de veículos '
        conta_credito = '1.07.04.12.01'
        parcelas = 60
    elif tipo == lista_tipo[1]:
        tipo = 0.1
        descricao = 'Depreciação de móveis '
        conta_credito = '1.07.04.12.02'
        parcelas = 120
    elif tipo == lista_tipo[2]:
        tipo = 0.2
        descricao = 'Depreciação de equipamentos '
        conta_credito = '1.07.04.12.03'
        parcelas = 60
    elif tipo == lista_tipo[3]:
        tipo = 0.04
        descricao = 'Depreciação de imóveis '
        conta_credito = '1.07.04.12.06'
        parcelas = 240
    elif tipo == lista_tipo[4]:
        tipo = 0.1
        descricao = 'Depreciação de instalações '
        conta_credito = '1.07.04.12.04'
        parcelas = 120
    data = entrada_data.get()
    button_create.pack()
    return valor, parcelas, tipo, data, conta_credito, descricao

#Função de save as
def set_save_destination():
    global save_spot
    save_spot = filedialog.askdirectory()
    save_spot = str(save_spot)

#Tela de sucesso
def sucess_window():
    window = Toplevel()
    window.geometry('220x75')
    root.resizable(width=False, height=False)

    sucess_label = Label(window, text='Planilha criada com sucesso!', font=('Helvetica, 12'), fg='green')
    sucess_label.pack()

    close_button = Button(window, text='Fechar', command=window.destroy, bg='#eb4034')
    close_button.pack()

#Definindo os entries
entrada_valor = Entry(root, width=50, bg='#dedede')
entrada_tipo = OptionMenu(root, variable, *lista_tipo)
entrada_tipo.config(bg='#dedede', width=30)
entrada_data = Entry(root, width=50, bg='#dedede')

#Definindo os label
text_valor = Label(root,text='Insira o valor da depreciação', font=("Helvetica", 8))
text_tipo = Label(root,text='Selectione o tipo de depreciação', font=("Helvetica", 8))
text_data = Label(root,text='Insira a data. Exemplo: 01/01/2022', font=("Helvetica", 8))

#Definindo os botões
send_button = Button(root, text='Confirmar', command=send_entries,height=2 , bg='#6b9ff2')
button_create = Button(root, text='Criar planilha', command= lambda: [set_save_destination(), criar_planilha(), sucess_window()], height=2 , bg='#6b9ff2')

#Posicionando elementos na tela
text_valor.pack(anchor='w', pady=3)
entrada_valor.pack(anchor='w', ipady=3)
text_data.pack(anchor='w', pady=3)
entrada_data.pack(anchor='w', ipady=3)
text_tipo.pack(anchor='w', pady=3)
entrada_tipo.pack(anchor='w')
send_button.pack(pady=10, ipadx=3)

#Chamando o Loop Principal do App
root.mainloop()